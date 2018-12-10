from openpyxl import Workbook, load_workbook
from models import *
from app import db


def import_file(file):
    wb = load_workbook(file.filename)
    # wb = load_workbook('2018-12-01-sales.xlsx')
    ws = wb.active

    sale = []
    for row in ws.iter_rows(min_row=1):
        if row[0].value == 'Date':
            continue
        for cell in row:
            # if cell.value == "":
            #     sale.append('NULL')
            # else:
            sale.append(cell.value)
        try:
            insert = Sales(date=sale[0], parts=sale[1], car_model=sale[2],
                           item_count=sale[3], value=sale[4])
            db.session.add(insert)
            db.session.commit()
            sale = []
        except Exception as e:
            print(e)
